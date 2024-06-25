import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Function to scrape data from an individual issue page
def scrape_issue_page(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    data = {}

    # Extracting relevant information
    data['issue_name'] = soup.find('h1', class_='page-header__title').text.strip()

    release_date = soup.find(string='Release Date')
    data['release_date'] = release_date.find_next('div').text.strip() if release_date else 'N/A'

    cover_date = soup.find(string='Cover Date')
    data['cover_date'] = cover_date.find_next('div').text.strip() if cover_date else 'N/A'

    writers = soup.find(string='Writer(s)')
    if writers:
        writer_elements = writers.find_next('div').find_all('a')
        data['writer'] = ', '.join(writer.text.strip() for writer in writer_elements)
    else:
        data['writer'] = 'N/A'

    artists = soup.find(string='Penciler(s)')
    if artists:
        artist_elements = artists.find_next('div').find_all('a')
        data['artist'] = ', '.join(artist.text.strip() for artist in artist_elements)
    else:
        data['artist'] = 'N/A'

    editors = soup.find(string='Editor(s)')
    if editors:
        editor_elements = editors.find_next('div').find_all('a')
        data['editor'] = ', '.join(editor.text.strip() for editor in editor_elements)
    else:
        data['editor'] = 'N/A'

    # Find all h2 elements with the specified class
    h2_elements = soup.find_all('h2', class_='pi-item pi-header pi-secondary-font pi-item-spacing pi-secondary-background')
    
    # Check if there are at least two such elements
    if len(h2_elements) >= 2:
        # First check the third element
        if len(h2_elements) >= 3:
            data['title'] = h2_elements[2].text.strip()  # Get the third element text
        else:
            data['title'] = h2_elements[1].text.strip()  # Get the second element text if third not found
    else:
        data['title'] = 'N/A'  # If less than two elements found, set title to N/A

    solicit_synopsis = soup.find(string='Solicit Synopsis')
    if solicit_synopsis:
        synopsis_text = solicit_synopsis.find_next('div').text.strip()
        data['solicit_synopsis'] = ' '.join(synopsis_text.split())
    else:
        data['solicit_synopsis'] = 'N/A'

    return data

# Function to update Excel sheet with the scraped data
def update_excel_sheet(filename, data):
    try:
        workbook = load_workbook(filename)
    except FileNotFoundError:
        workbook = Workbook()
    
    sheet = workbook.active

    headers = ['Issue Name', 'Release Date', 'Cover Date', 'Writer', 'Artist', 'Editor', 'Title', 'Solicit Synopsis']
    
    # Check if headers are already in the sheet
    if sheet['A1'].value != headers[0]:
        sheet.append(headers)  # Add headers if they are not already in the sheet

    row = [data.get(header.lower().replace(' ', '_'), '') for header in headers]
    sheet.append(row)

    workbook.save(filename)

# Main function to orchestrate the scraping and updating Excel sheet
def main():
    issue_url = input("Enter the URL of the individual issue page: ")
    excel_filename = 'Marvel_Comics_Issue_Data.xlsx'

    scraped_data = scrape_issue_page(issue_url)

    update_excel_sheet(excel_filename, scraped_data)
    print(f"Data has been successfully scraped and saved to {excel_filename}")

if __name__ == '__main__':
    main()
